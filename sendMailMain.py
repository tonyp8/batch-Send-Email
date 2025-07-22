# -*- coding: UTF-8 -*-
from os import listdir,getcwd,popen,remove
#import os
import time
import json
import openpyxl
import threading
from numpy import array_split 

import sendMailCore
#import batchChangeNamePyperclip

#content = getcwd()
#print(listdir(content))
def findConfig(forcedReconstruction = False):
    '''
    查找配置文件是否存在,用于初启动与重置设置
    Args:
        #files:根目录文件列表
        forcedReconstruction:是否重置设置(bool)
    '''
    if (not('config.json' in listdir(getcwd()))) or forcedReconstruction:
        try:
            remove(r'''config.json''')
        except FileNotFoundError:
            print('[WARNING] 未找到配置文件!')
        #system("copy config.cfg config.json")
        popen("copy config.cfg config.json")
        time.sleep(0.5)
        print('[INFO] 配置文件已生成!')

def readConfig():
    '''#读取配置文件'''
    try:
        with open(r'''config.json''', 'r',encoding="utf-8") as f:
            data = json.load(f)
        #print(data)
        print('-'*10,'Start','-'*10)
        print('[INFO] 配置文件读取成功')
    except Exception as e:
        print(f"[ERROR] 读取配置文件错误:{e}")

    '''
    #data = {'a':'b'}
    a = data.pop('//','')
    b = data.pop('settings','')
    c = data.pop('exampleAccount','')
    if not (a and b and c):
        print('[ERROR] 读取配置文件出错!请按照指引修改或重置!')
    '''
    
    #print(list(data.keys()))
    return data


def readSheet(sheetContent = 'expertSheet.xlsx'):
    '''读取达人列表
    Args:
        sheetContent:列表位置
    '''
    #global workbook
    sheet = ''
    for item in listdir(getcwd()):
        if '~$' in item:
            print('[WARNING] 请保存并关闭目录下所有excel表!')
    try:
        workbook = openpyxl.load_workbook(sheetContent)  # 返回一个workbook数据类型的值
        sheet = workbook.active #获取活动工作表
        print('[INFO] 达人列表读取成功!')
        
    except Exception as e:
        if str(e) == f"[Errno 2] No such file or directory: '{sheetContent}'":
            print('[WARNING] 未找到表格!')
            popen("copy DefaultSheet.sheet expertSheet.xlsx")
            time.sleep(0.5)
            print("[INFO] 默认表格'expertSheet.xlsx'已生成!")
            sheet = readSheet()
        else:
            print('[ERROR] 读取表格时发生错误!',e)
    return sheet,workbook

    
def getReciver(sheet,key,chargerNameSetting):
    '''
    获取收信人
    Args:
        sheet:达人建联表
        
        key:字典,xlsx关键读取位置(列),json如下
            {"//":"负责人",
            "responser":3,
            "//":"达人ID",
            "expertID":4,
            "//":"达人邮箱地址",
            "expertEmail":9,
            "//":"邮件发送情况(bool)",
            "sendingStatus":15}
            
        chargerNameSetting:字典,发送人名称匹配设定,json如下
           {"//":"设置选项:0.忽略大小写完全匹配;1.排除",
            "statue":0,
            "senderName":"coco"}
    Return:
        字典，样式如下：
        {'达人名':{'mail':'abc@def.com','region':'DE','row':1}} row作为标识状态使用
    '''
    reciverdict = {} #收件人:邮箱
    mode = {0:'忽略大小写完全匹配',1:'排除'}
    #print(chargerNameSetting)
    print(f'''[INFO] 当前筛选模式:{chargerNameSetting['statue']},{mode[chargerNameSetting['statue']]}''')
    resName = chargerNameSetting['chargerName'].upper()# 负责人名字
    print(f'''[INFO] 匹配/排除名称: {resName}''')
    
    for trow in range(1,sheet.max_row+1): #忽略首行
        if not sheet.cell(row = trow,column = key['sendingStatus']).value: #没有发送
            if ((not chargerNameSetting['statue']) and resName == sheet.cell(row = trow,column = key['responser']).value) or (chargerNameSetting['statue'] and resName != sheet.cell(row = trow,column = key['responser']).value):
                #模式0，负责人名称相等 / 模式1 负责名字不等
                
                expertName = sheet.cell(row = trow,column = key['expertID']).value
                expertRegion = sheet.cell(row = trow,column = key['expertRegion']).value                    
                expertEmail = sheet.cell(row = trow,column = key['expertEmail']).value

                if expertEmail and '@' in expertEmail: #忽略没有邮箱的
                    #添加筛选结果
                    reciverdict[expertName] = {'mail':expertEmail,'region':expertRegion,'row':trow}
                else:
                    print(f'[WARNING] {expertName} 没有邮箱,已排除')                 
    return reciverdict

def split_dict_avg(d, n):
    '''dict,份数n'''
    items = list(d.items())
    k, m = divmod(len(items), n)
    return [dict(items[i * k + min(i, m):(i + 1) * k + min(i + 1, m)]) for i in range(n)]



if __name__ == "__main__":
    #获取运行目录文件
    #files = listdir(getcwd())
    #print(files)
    #读取配置文件
    findConfig()
    config = readConfig()
    #print(config)
    sheet,workbook = readSheet(config['settings']['content']) #此处workbook无用处
    #print(config['accounts']['bili'])
    reciverdict = getReciver(sheet,config['settings']["xlsxKeys"],config['settings']["chargerName"])
    #print(reciverdict)
    '''
    message = {'subject':'批发测试',
            'content':f'找到以下达人\n{reciverdict}',
            'attachment':[]}
            
    #下面为批量不改名发送！'''
    
    reciver = []
    for name in list(reciverdict.keys()):
        reciver.append(reciverdict[name]['mail'])
    #reciver = ['bilicoco29@gmail.com']
    print(f'\n[INFO] 邮件即将发送到{reciver}！')

    if not bool(reciver):
        print('[WARNING] 没有任何需要发送的邮箱，请手动关闭软件！')
    
    #发送前等待
    to=10
    print(f'[WARNING] {to}s内按Ctrl+C退出软件取消发送！')
    for t in range(to,0,-1):
        print(t)
        time.sleep(1)

    del config['accounts']['//']
    del config['accounts']['exampleAccount']
    del config['settings']['mailModelContent']['//']
    
    if len(config['accounts']) == 0:
        print('[ERROR] 没有配置任何邮箱账号!')
    elif len(config['accounts']) == 1: #仅配置1个账号
        sendMailCore.send_emails(list(config['accounts'].keys())[0],list(config['accounts'].values())[0],reciverdict,config['settings'])
    else:
        if config['settings']['singleSender']['enabled']:
            defaultName = config['settings']['singleSender']['defaultSenderName']
            print('[INFO] 单邮箱发送')
            if defaultName in config['accounts']:
                #单线程,仅使用bili号
                sendMailCore.send_emails(defaultName,config['accounts'][defaultName],reciverdict,config['settings'])
            else:
                print('[WARNING] 单邮箱发送设置默认账户有误,已自动选择第一个发送')
                sendMailCore.send_emails(list(config['accounts'].keys())[0],list(config['accounts'].values())[0],reciverdict,config['settings'])
        else:
            print('[INFO] 多邮箱发送')
            recDictList = split_dict_avg(reciverdict,len(config['accounts']))
            #print(recDictList,len(recDictList),len(config['accounts']))
            # 创建停止用的事件对象
            stop_event = threading.Event()
            #print(config['accounts'])
            threads = []
            for i,account in enumerate(list(config['accounts'].keys())):
                print(f'[INFO] 线程{i}({account})创建中')
                thread = threading.Thread(target = sendMailCore.send_emails,
                                          args=(account,config['accounts'][account],recDictList[i],),
                                          kwargs={'settings':config['settings'],
                                                  'threadID':account,
                                                  'stop_event':stop_event})
                thread.start()
                threads.append(thread)
                time.sleep(0.1)
                print(f'[INFO] 线程{i}({account})已创建')
                time.sleep(config['settings']["intervalSendingTime"]//len(config['accounts']))
                if config['settings']["staggeredSending"]: #关闭交错发送
                    thread.join()
            #stop_event.set()
            #stop_threads = True
            for t in threads:
                t.join()
                #print(t)
            print('发送完成！')
            time.sleep(1800)
